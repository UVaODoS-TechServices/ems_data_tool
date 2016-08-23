# -*- coding: utf-8 -*-

""" Updates ODOS EMS system using data from @UVa and external sources. """

import json
import os
import socket
import sys
import time

from argparse import ArgumentParser
from ConfigParser import SafeConfigParser
from multiprocessing import Process, Queue
from tempfile import NamedTemporaryFile
from requests import Session

import openpyxl
import pyodbc

from smb.SMBConnection import SMBConnection
from unidecode import unidecode

from lib.atuva import get_items
from lib.groups import process_departments, process_organizations
from lib.tools import create_connmsg, parse_spreadsheet, unpack

def fetch_departments(config):
    """ Fetches departments from network storage. """

    for option, filename in config.items("MAIN"):
        if "_source" in option:
            file_obj = NamedTemporaryFile()
            conn = SMBConnection(username=config.get("MAIN", "username"),
                                 password=config.get("MAIN", "password"),
                                 my_name=socket.gethostname(),
                                 domain=config.get("MAIN", "domain"),
                                 remote_name=config.get("MAIN", "sharedrive"),
                                 is_direct_tcp=True)
            conn.connect(socket.gethostbyname(config.get("MAIN", "sharedrive")), 445)

            print "fetching %s via smb..." % (os.path.basename(filename))

            _ = conn.retrieveFile(config.get("MAIN", "service"), filename, file_obj)
            data = parse_spreadsheet(file_obj)

            conn.close()
            file_obj.close()

            for datum in data:
                yield datum


def get_memberships(organization, config, session=None):
    """ Fetches memberships of for each position for the organization. """

    params = {"organizationId": organization['organizationId'], \
              "currentMembershipsOnly": True}

    for position in ("Event Requester 1", "Event Requester 2"):
        params['positionTemplateName'] = position

        items = get_items(config, "memberships", params, session)

        for item in items:
            yield item


class WritingWorker(object):
    """ Object for writing data in a process. """

    def __init__(self, filename):
        self.filename = filename
        self.memo = []
        self.state = 0
        self.file_obj = None

    def run(self, queue):
        """ Called during start of loop in process. """

        self.file_obj = open(self.filename, 'w')
        self.state = 1

        while self.state:
            data = queue.get()

            if isinstance(data, dict):
                for item in unpack(data):
                    seen = (item[0], item[-3])

                    if seen not in self.memo:
                        self.memo.append(seen)

                    else:
                        continue

                    self.file_obj.write(json.dumps(item) + '\n')

            elif isinstance(data, str):
                if data == "STOP":
                    self.file_obj.close()
                    self.state = 0

                return

            time.sleep(0.1)


def update_database(config, filename):
    """ Helper function to update out-of-band. """

    groups = []
    with open(filename, 'r') as fin:
        for line in fin:
            groups.append(json.loads(line.strip()))

    connmsg = create_connmsg(
        svr=config.get("DATABASE", "server"),
        drv=config.get("DATABASE", "driver"),
        db=config.get("DATABASE", "database"),
        un=config.get("DATABASE", "username"),
        pwd=config.get("DATABASE", "password"),
        tc=config.get("DATABASE", "trusted_connection")
        )

    conn = pyodbc.connect(connmsg)
    cur = conn.cursor()

    query = "DELETE FROM EMS_Staging.dbo.tblPeople;"

    cur.execute(query)

    debuglog = open("debuglog.txt", 'w')

    query = "INSERT INTO EMS_Staging.dbo.tblPeople(PersonnelNumber, FirstName, "
    query += "LastName, Title, MiddleInitial, EMailAddress, Phone, Fax, Address1, "
    query += "Address2, City, State, ZipCode, Country, NetworkID, GroupID, GroupName, "
    query += "GroupType) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);"

    for group in groups:
        try:
            cur.execute(query, *group)
            cur.commit()

        except:
            debuglog.write(json.dumps(group) + '\n')


def get_staged_users(config, filename):
    """ Fetches and dumps users currently in staging table to file. """

    connmsg = create_connmsg(
        svr=config.get("DATABASE", "server"),
        drv=config.get("DATABASE", "driver"),
        db=config.get("DATABASE", "database"),
        un=config.get("DATABASE", "username"),
        pwd=config.get("DATABASE", "password"),
        tc=config.get("DATABASE", "trusted_connection")
        )

    conn = pyodbc.connect(connmsg)
    cur = conn.cursor()

    query = "SELECT PersonnelNumber, FirstName, LastName, Title, MiddleInitial, "
    query += "EmailAddress, Phone, Fax, Address1, Address2, City, State, ZipCode, "
    query += "Country, NetworkID, GroupID, GroupName, GroupType FROM EMS_Staging.dbo.tblPeople;"

    cur.execute(query)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Staged Users"

    result = cur.fetchone()
    results = []

    while result is not None and result != '':
        results.append([field for field in result])
        result = cur.fetchone()

    headers = config.get("KEYS", "group_fields").split(', ')

    for header in headers:
        worksheet.cell(row=1, column=(headers.index(header) + 1)).value = header

    for result in results:
        r_index = results.index(result) + 2

        for field in result:
            f_index = result.index(field) + 1

            if field is None:
                worksheet.cell(row=r_index, column=f_index).value = ""

            else:
                try:
                    worksheet.cell(row=r_index, column=f_index).value = field

                except UnicodeDecodeError:
                    value = unidecode(unidecode(field).decode("utf-8"))
                    worksheet.cell(row=r_index, column=f_index).value = value

    workbook.save(filename)


def verify_database(config, filenames):
    """
        Runs scripts against database connection to determine validity.
        Scripts must return results only on verfication failure.
        Requires ConfigParser style config.
        Accepts ConfigParser style config and list of filenames for script files.
        Generates filename, results for each script run.
    """

    if 'win' in sys.platform:
        endings = '\n\n'

    else:
        endings = '\r\n\r\n'

    connmsg = create_connmsg(
        svr=config.get("DATABASE", "server"),
        drv=config.get("DATABASE", "driver"),
        db=config.get("DATABASE", "database"),
        un=config.get("DATABASE", "username"),
        pwd=config.get("DATABASE", "password"),
        tc=config.get("DATABASE", "trusted_connection")
        )

    conn = pyodbc.connect(connmsg)
    cur = conn.cursor()

    retval = []

    for filename in filenames:
        if not os.path.exists(filename):
            message = "%s cannot be read, bailing!" % (filename,)
            raise IOError(message)

        with open(filename, 'r') as fin:
            data = fin.read().split(endings)

        for datum in data:
            cur.execute(datum)

        results = []
        result = cur.fetchone()

        while result is not None and result != '':
            results.append([field for field in result])
            result = cur.fetchone()

        retval.extend(results)

    if len(results) > 0:
        with open("errors.txt", 'w') as fout:
            for result in results:
                if isinstance(result, str):
                    fout.write(result)

                else:
                    print result
                    sys.exit(-1)
                    fout.write(json.dumps(result))

        return False

    return True


def prune_failed(departments):
    """ Removes empty rows from the departments list and returns result. """

    good = []

    for department in departments:
        if not isinstance(department, dict):
            continue

        if any((department[key] is None for key in department.keys())):
            continue

        good.append(department)

    return good


def main():
    """ Where the magic happens. """

    parser = ArgumentParser()
    parser.add_argument(
        "-o",
        "--outfile",
        nargs='?',
        help="filename to output data to (i.e. somefile.json)",
        required=True
        )
    parser.add_argument(
        "-c",
        "--configfile",
        nargs='?',
        help="filename that contains configuration data (i.e. somefile.ini)",
        required=True
        )

    args = parser.parse_args()
    filename = args.outfile

    print "using file {fname} as output...".format(fname=filename)

    config = SafeConfigParser()
    config.optionxform(str())
    config.read(args.configfile)

    session = Session()

    params = {"status": "Active",
              "type": "Contracted Independent Organization",
              "page": 1,
              "pageSize": 100}

    print "starting..."

    worker = WritingWorker(filename)
    queue = Queue()
    process = Process(target=worker.run, args=(queue,))
    process.daemon = True
    process.start()

    for organization in get_items(config, "organizations", params, session):
        memberships = list(get_memberships(
            organization=organization,
            config=config,
            session=session
            ))

        queue.put(process_organizations(([organization,], memberships), config))

    departments = list(fetch_departments(config))
    departments = prune_failed(departments)

    for department in departments:
        queue.put(process_departments([department], config))

    queue.put("STOP")
    process.join()

    sys.exit(0)


if __name__ == "__main__":
    if "win" in sys.platform:
        from multiprocessing import freeze_support

    main()

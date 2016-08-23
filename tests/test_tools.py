# -*- coding: utf-8 -*-

""" Unittests for tools library. """

import sys
import time
import unittest

import StringIO

import openpyxl

from lib.tools import StatusMessage, parse_spreadsheet, Timer, pack, unpack

class TestTools(unittest.TestCase):
    """ Tests for validating tools library. """

    def test_status_message(self):
        """ Tests status message function. """

        s_msg = StatusMessage()

        self.assertEqual(s_msg("test"), "test")
        self.assertEqual(s_msg("test"), None)

    def test_parse_spreadsheet(self):
        """ Tests ability to parse spreadsheets with headers. """

        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        #fudge a worksheet
        worksheet.cell(row=1, column=1).value = "Column A"
        worksheet.cell(row=1, column=2).value = "Column B"
        worksheet.cell(row=2, column=1).value = "Value A"
        worksheet.cell(row=2, column=2).value = "Value B"

        file_obj = StringIO.StringIO() #write to memory
        workbook.save(file_obj)

        if file_obj.tell() > 0:
            file_obj.seek(0)

        result = parse_spreadsheet(file_obj)

        self.assertIsInstance(result, list)
        self.assertIsInstance(result[0], dict)

    def test_timer(self):
        """ Tests timer decorator. """

        dummyio = StringIO.StringIO()
        sys.stdout = dummyio #redirect stdout to capture output

        @Timer
        def wait(secs):
            """ Waits for specified number of seconds. """

            time.sleep(secs)

        sys.stdout = sys.__stdout__ #restore stdout

        if dummyio.tell() > 0:
            dummyio.seek(0)

        result = dummyio.read()

        self.assertIsInstance(result, str)

    def test_unpack(self):
        """ Tests ability to unpack processed groups. """
        #fudge valid item
        requester = {"FirstName": "CAFE", "LastName": "BABE", \
                     "MiddleInitial": "", "EMailAddress": "0DD@CAFE.C0", \
                     "Phone": "", "Fax": "", "Address1": "", "Address2": "", \
                     "City": "", "State": "", "ZipCode": "", "Country": "", \
                     "NetworkID": "0DD"}
        organization = {"GroupName": "FACED00D", "GroupType": "DEADBEEF", \
                        "Event Requester 1": requester}
        item = {"0DDC0FFEE": organization}

        result = list(unpack(item))

        self.assertIsInstance(result, list)

    def test_pack(self):
        """ Tests ability to pack/repack processed groups. """

        requester = ["0DD", "CAFE", "BABE", "Event Requester 1", "", "0DD@CAFE.C0", "", "", "", "", "", "", "", "", "0DD", "0DDC0FFEE", "FACED00D", "DEADBEEF"]
        result = list(pack([requester]))

        self.assertIsInstance(result[0], dict)


if __name__ == "__main__":
    pass
